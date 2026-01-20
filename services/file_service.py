"""
File Service - Handles file upload validation and Excel processing

This service encapsulates all file handling, validation, and Excel
data extraction logic.
"""

from typing import Tuple, Dict, List
from werkzeug.datastructures import FileStorage
import openpyxl
from io import BytesIO


class FileService:
    """Service for handling file uploads and Excel processing."""

    def __init__(self, upload_extensions: List[str], max_content_length: int):
        """
        Initialize the file service.

        Args:
            upload_extensions: List of allowed file extensions (e.g., ['.xlsx', '.xls'])
            max_content_length: Maximum file size in bytes
        """
        self.upload_extensions = upload_extensions
        self.max_content_length = max_content_length

    def validate_uploaded_file(self, file: FileStorage) -> Tuple[bool, str]:
        """
        Comprehensive file upload validation.

        Args:
            file: FileStorage object from Flask request

        Returns:
            Tuple of (is_valid, error_message)
            error_message is None if valid
        """
        if not file or not file.filename:
            return False, "No file selected"

        filename = file.filename.lower().strip()

        # Check for path traversal attempts
        if '/' in filename or '\\' in filename or '..' in filename:
            return False, "Invalid filename"

        # Check file extension
        if not any(filename.endswith(ext) for ext in self.upload_extensions):
            return False, "Invalid file type. Only .xlsx and .xls allowed"

        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning

        if file_size > self.max_content_length:
            return False, "File too large. Maximum 10MB"

        if file_size == 0:
            return False, "File is empty"

        # Validate file magic bytes
        header = file.read(8)
        file.seek(0)

        # Check for valid Excel file signatures
        is_xlsx = header[:2] == b'PK'  # ZIP format (xlsx)
        is_xls = header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'  # OLE format (xls)

        if not (is_xlsx or is_xls):
            return False, "File is not a valid Excel file"

        return True, None

    def process_excel_upload(self, file: FileStorage) -> Tuple[Dict[str, List[Dict]], List[str]]:
        """
        Process uploaded Excel file and extract vacancy data.

        Expected columns:
        1. Vacancy Name
        2. Recruiter Name
        3. Role Type (Easy/Medium/Hard)
        4. Internal? (Yes/No)
        5. Stage (Sourcing/Screening/Interview/Offer/Pre-Hire Checks or blank)

        Args:
            file: FileStorage object from Flask request.files

        Returns:
            Tuple of (recruiters_dict, errors_list)
            recruiters_dict: {recruiter_name: [vacancy_dicts]}
            errors_list: List of error messages
        """
        errors = []
        recruiters_dict = {}

        try:
            # Read Excel file
            file_bytes = file.read()
            workbook = openpyxl.load_workbook(BytesIO(file_bytes))
            sheet = workbook.active

            # Check if file has data
            if sheet.max_row < 2:
                errors.append("Excel file is empty or has no data rows.")
                return recruiters_dict, errors

            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]

            # Validate required columns (case-insensitive)
            required_columns = ['vacancy name', 'recruiter name', 'role type', 'internal?', 'stage']
            headers_lower = [str(h).lower().strip() if h else '' for h in headers]

            missing_columns = []
            for req_col in required_columns:
                if req_col not in headers_lower:
                    missing_columns.append(req_col)

            if missing_columns:
                errors.append(f"Missing required columns: {', '.join(missing_columns)}")
                return recruiters_dict, errors

            # Get column indices
            col_indices = {}
            for req_col in required_columns:
                try:
                    col_indices[req_col] = headers_lower.index(req_col)
                except ValueError:
                    pass

            # Process data rows
            for row_num in range(2, sheet.max_row + 1):
                row = sheet[row_num]

                # Extract values
                vacancy_name = row[col_indices['vacancy name']].value
                recruiter_name = row[col_indices['recruiter name']].value
                role_type = row[col_indices['role type']].value
                internal_str = row[col_indices['internal?']].value
                stage = row[col_indices['stage']].value

                # Validate recruiter name
                if not recruiter_name or str(recruiter_name).strip() == '':
                    errors.append(f"Row {row_num}: Empty Recruiter Name")
                    continue

                recruiter_name = str(recruiter_name).strip()

                # Validate vacancy name
                if not vacancy_name or str(vacancy_name).strip() == '':
                    vacancy_name = f"Vacancy {row_num - 1}"

                vacancy_name = str(vacancy_name).strip()

                # Validate role type
                if not role_type or str(role_type).strip() == '':
                    errors.append(f"Row {row_num}: Empty Role Type for {vacancy_name}")
                    continue

                role_type = str(role_type).strip().lower()
                if role_type not in ['easy', 'medium', 'hard']:
                    errors.append(f"Row {row_num}: Invalid Role Type '{role_type}' for {vacancy_name}. Must be Easy, Medium, or Hard.")
                    continue

                # Validate internal
                if not internal_str or str(internal_str).strip() == '':
                    internal_str = 'No'

                internal_str = str(internal_str).strip().lower()
                if internal_str not in ['yes', 'no']:
                    errors.append(f"Row {row_num}: Invalid Internal value '{internal_str}' for {vacancy_name}. Must be Yes or No.")
                    continue

                is_internal = (internal_str == 'yes')

                # Validate stage
                if not stage or str(stage).strip() == '':
                    stage = ''
                else:
                    stage = str(stage).strip().lower()
                    valid_stages = ['sourcing', 'screening', 'interview', 'offer', 'pre-hire checks', '']
                    if stage not in valid_stages:
                        errors.append(f"Row {row_num}: Invalid Stage '{stage}' for {vacancy_name}.")
                        continue

                # Add to recruiters dict
                if recruiter_name not in recruiters_dict:
                    recruiters_dict[recruiter_name] = []

                recruiters_dict[recruiter_name].append({
                    'vacancy_name': vacancy_name,
                    'role_type': role_type,
                    'is_internal': is_internal,
                    'stage': stage
                })

        except openpyxl.utils.exceptions.InvalidFileException:
            errors.append("Invalid file format. Please upload a valid Excel file (.xlsx or .xls).")
        except Exception as e:
            errors.append(f"Error processing Excel file: {str(e)}")

        return recruiters_dict, errors

    @staticmethod
    def generate_capacity_template() -> openpyxl.Workbook:
        """
        Generate a sample Excel template for capacity tracker.

        Returns:
            openpyxl Workbook object ready to be saved
        """
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Capacity Data"

        # Add headers
        headers = ['Vacancy Name', 'Recruiter Name', 'Role Type', 'Internal?', 'Stage']
        sheet.append(headers)

        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add sample data
        sample_data = [
            ['Senior Developer', 'John Smith', 'Hard', 'No', 'Screening'],
            ['Marketing Assistant', 'Jane Doe', 'Easy', 'Yes', 'Sourcing'],
            ['Finance Manager', 'John Smith', 'Hard', 'No', ''],
            ['HR Coordinator', 'Jane Doe', 'Medium', 'No', 'Interview'],
            ['IT Support', 'Bob Wilson', 'Easy', 'Yes', 'Offer'],
        ]

        for row in sample_data:
            sheet.append(row)

        # Adjust column widths
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 18

        # Add instructions sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            ['Recruitment Capacity Tracker - Template Instructions'],
            [''],
            ['Column Requirements:'],
            ['1. Vacancy Name: Name of the vacancy (optional, will auto-generate if blank)'],
            ['2. Recruiter Name: Name of the recruiter (required)'],
            ['3. Role Type: Easy, Medium, or Hard (required, case-insensitive)'],
            ['4. Internal?: Yes or No (required, case-insensitive)'],
            ['5. Stage: Sourcing, Screening, Interview, Offer, Pre-Hire Checks, or blank (optional)'],
            [''],
            ['Business Rules:'],
            ['- Easy roles: Max 30 at full capacity (3.33% each)'],
            ['- Medium roles: Max 20 at full capacity (5% each)'],
            ['- Hard roles: Max 12 at full capacity (8.33% each)'],
            ['- Internal roles take 75% less time (0.25 multiplier)'],
            ['- Stage weights: Sourcing (20%), Screening (40%), Interview (20%), Offer (10%), Pre-Hire (10%), None (100%)'],
            [''],
            ['Example Calculations:'],
            ['- External, Easy, No Stage: 1/30 = 3.33%'],
            ['- Internal, Hard, Screening: (1/12) × 0.25 × 0.4 = 0.83%'],
            ['- External, Medium, Interview: (1/20) × 0.2 = 1%'],
        ]

        for row in instructions:
            instructions_sheet.append(row)

        return workbook
