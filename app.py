"""
Newton's Repository - A collection of projects, stories, and experiments
"""
from flask import Flask, render_template, abort, url_for, redirect, request
from datetime import datetime
from pathlib import Path
import os
import re
from config import get_config
from models import Article, BlogCategory, RoleType, RecruitmentStage, Vacancy, Recruiter

# Load configuration based on environment
app = Flask(__name__)
app.config.from_object(get_config())

@app.after_request
def set_security_headers(response):
    """Apply security headers to all responses."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'

    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "img-src 'self' data: https:; "
        "font-src 'self' https://cdn.jsdelivr.net; "
        "connect-src 'self'; "
        "frame-ancestors 'self'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )

    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'

    return response

# Initialize CSRF Protection
from flask_wtf.csrf import CSRFProtect
csrf = CSRFProtect(app)

# Raw category data for initialization
_RAW_CATEGORIES = {
    "morecambe-fm26": {
        "id": "morecambe-fm26",
        "name": "Morecambe FC",
        "subtitle": "FM26 Save",
        "description": "Following Morecambe FC from administration to glory.",
        "image": "morecambe-logo.png",
        "articles": [
            {"id": "the-journey-begins", "title": "How did we fall so far?", "date": "2024-01-15", "filename": "article1.txt", "part": 1},
            {"id": "first-season-struggles", "title": "The struggle is real, or is it?", "date": "2024-02-20", "filename": "article2.txt", "part": 2},
            {"id": "transfer-window-rebuild", "title": "Crossing the line", "date": "2024-03-10", "filename": "article3.txt", "part": 3},
            {"id": "turning-point", "title": "Out with the old and in with the older", "date": "2024-04-05", "filename": "article4.txt", "part": 4},
            {"id": "promotion-push", "title": "The Crucible", "date": "2024-05-12", "filename": "article5.txt", "part": 5},
            {"id": "glory-day", "title": "Disaster and Triumph", "date": "2024-06-01", "filename": "article6.txt", "part": 6},
        ]
    }
}

# Convert raw data to typed BlogCategory objects
BLOG_CATEGORIES = {}
for cat_id, data in _RAW_CATEGORIES.items():
    articles = [Article(**article, category_id=cat_id) for article in data['articles']]
    BLOG_CATEGORIES[cat_id] = BlogCategory(
        id=data['id'],
        name=data['name'],
        subtitle=data['subtitle'],
        description=data['description'],
        image=data['image'],
        articles=articles
    )

PROJECTS = [
    {"id": "capacity-tracker", "name": "Recruitment Capacity Tracker", "description": "Calculate team capacity for recruitment workloads.", "status": "active", "url": "/projects/capacity-tracker"},
    {"id": "meal-generator", "name": "Meal Generator", "description": "A tool to help plan weekly meals.", "status": "planned"},
    {"id": "fm-tools", "name": "FM Analytics Tools", "description": "Utilities for analyzing FM save data.", "status": "planned"},
]

def get_article_content(filename):
    """Safely read article content with path validation."""
    # Validate filename - only allow alphanumeric, dash, underscore, and dot
    if not filename or not re.match(r'^[\w\-\.]+$', filename):
        return None

    # Use centralized path configuration
    articles_dir = app.config['ARTICLES_DIR']
    filepath = (articles_dir / filename).resolve()

    # Ensure the resolved path is still within articles directory (prevents traversal)
    try:
        filepath.relative_to(articles_dir)
    except ValueError:
        # Path is outside articles directory
        return None

    # Read file with error handling
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return f.read()
    except (FileNotFoundError, PermissionError):
        return None

def format_date(date_string):
    date_obj = datetime.strptime(date_string, "%Y-%m-%d")
    return date_obj.strftime("%B %d, %Y")

def calculate_reading_time(text):
    words = len(text.split())
    return max(1, round(words / 200))

def get_excerpt(text, sentence_count=2):
    sentences = re.split(r'(?<=[.!?])\s+', text.strip())
    excerpt = ' '.join(sentences[:sentence_count])
    if len(excerpt) > 200:
        excerpt = excerpt[:197] + '...'
    return excerpt

def parse_content(text):
    blocks = []
    for para in [p.strip() for p in text.split("\n\n") if p.strip()]:
        is_heading = re.match(r'^Part\s+\d+', para, re.IGNORECASE) or (len(para) < 80 and not para.endswith('.'))
        blocks.append({"type": "heading" if is_heading else "paragraph", "content": para})
    return blocks

def get_category_articles(category_id):
    """Get category and its articles with enriched data."""
    category = BLOG_CATEGORIES.get(category_id)
    if not category:
        return None, []

    # Enrich articles with content, reading time, and excerpt
    enriched_articles = []
    for article in category.articles:
        content = get_article_content(article.filename)
        # Create a copy with enriched data
        enriched = Article(
            id=article.id,
            title=article.title,
            date=article.date,
            filename=article.filename,
            part=article.part,
            category_id=article.category_id,
            content=content,
            reading_time=calculate_reading_time(content) if content else 0,
            excerpt=get_excerpt(content) if content else ""
        )
        enriched_articles.append(enriched)

    return category, sorted(enriched_articles, key=lambda x: x.part)

def get_article_by_id(category_id, article_id):
    """Get category and specific article by ID."""
    category = BLOG_CATEGORIES.get(category_id)
    if not category:
        return None, None

    article = category.get_article_by_id(article_id)
    return category, article

def get_prev_next_articles(category_id, current_part):
    """Get previous and next articles for navigation."""
    category = BLOG_CATEGORIES.get(category_id)
    if not category:
        return None, None

    prev_article, next_article = None, None
    for article in category.articles:
        if article.part == current_part - 1:
            prev_article = article
        elif article.part == current_part + 1:
            next_article = article

    return prev_article, next_article

def get_latest_article():
    """Get the most recent article across all categories."""
    latest, latest_date, latest_category = None, None, None

    for cat_id, category in BLOG_CATEGORIES.items():
        for article in category.articles:
            if latest_date is None or article.date_obj > latest_date:
                latest, latest_date, latest_category = article, article.date_obj, category

    if latest and latest_category:
        content = get_article_content(latest.filename)
        # Return enriched Article object
        enriched = Article(
            id=latest.id,
            title=latest.title,
            date=latest.date,
            filename=latest.filename,
            part=latest.part,
            category_id=latest_category.id,
            content=content,
            excerpt=get_excerpt(content) if content else "",
            reading_time=calculate_reading_time(content) if content else 0
        )
        # Add category name as attribute for template convenience
        enriched.category_name = latest_category.name
        return enriched

    return None

# ========== RECRUITMENT CAPACITY TRACKER FUNCTIONS ==========

# Stage-based time weighting multipliers
STAGE_MULTIPLIERS = {
    'sourcing': 0.2,
    'screening': 0.4,
    'interview': 0.2,
    'offer': 0.1,
    'pre-hire checks': 0.1,
    '': 1.0,  # No stage specified = full time
    'none': 1.0
}

# Base capacity values
BASE_CAPACITY = {
    'easy': 1/30,      # 3.33% per vacancy
    'medium': 1/20,    # 5% per vacancy
    'hard': 1/12       # 8.33% per vacancy
}

def validate_uploaded_file(file):
    """Comprehensive file upload validation."""
    if not file or not file.filename:
        return False, "No file selected"

    filename = file.filename.lower().strip()

    # Check for path traversal attempts
    if '/' in filename or '\\' in filename or '..' in filename:
        return False, "Invalid filename"

    # Check file extension
    if not any(filename.endswith(ext) for ext in app.config['UPLOAD_EXTENSIONS']):
        return False, "Invalid file type. Only .xlsx and .xls allowed"

    # Check file size
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning

    if file_size > app.config['MAX_CONTENT_LENGTH']:
        return False, "File too large. Maximum 10MB"

    if file_size == 0:
        return False, "File is empty"

    # Validate file magic bytes
    header = file.read(8)
    file.seek(0)

    # Check for valid Excel file signatures
    is_xlsx = header[:2] == b'PK'  # ZIP format (xlsx)
    is_xls = header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'  # OLE format (xls)

    if not (is_xlsx or is_xls):
        return False, "File is not a valid Excel file"

    return True, None

def calculate_vacancy_capacity(role_type, is_internal=False, stage=''):
    """
    Calculate capacity usage for a single vacancy with enhanced business logic.

    Business rules:
    - Easy: 1/30 = 3.33% base capacity
    - Medium: 1/20 = 5% base capacity
    - Hard: 1/12 = 8.33% base capacity
    - Internal roles: 0.25 multiplier (75% less time)
    - Stage-based: Sourcing (20%), Screening (40%), Interview (20%),
                   Offer (10%), Pre-Hire Checks (10%), None (100%)

    Formula: base_capacity × internal_multiplier × stage_multiplier

    Args:
        role_type (str): 'easy', 'medium', or 'hard'
        is_internal (bool): True if internal-only role
        stage (str): Recruitment stage or empty string

    Returns:
        float: Capacity used by this vacancy (0-1 scale)
    """
    # Get base capacity
    role_type_lower = role_type.lower()
    if role_type_lower not in BASE_CAPACITY:
        raise ValueError(f"Invalid role type: {role_type}")

    base_capacity = BASE_CAPACITY[role_type_lower]

    # Apply internal multiplier
    internal_multiplier = 0.25 if is_internal else 1.0

    # Apply stage multiplier
    stage_lower = stage.lower().strip()
    stage_multiplier = STAGE_MULTIPLIERS.get(stage_lower, 1.0)

    # Calculate final capacity
    vacancy_capacity = base_capacity * internal_multiplier * stage_multiplier

    return vacancy_capacity

def calculate_recruiter_capacity_from_vacancies(vacancies):
    """
    Calculate capacity for a recruiter from a list of individual vacancies.

    Args:
        vacancies (list): List of vacancy dicts with keys:
                         - vacancy_name
                         - role_type
                         - is_internal
                         - stage

    Returns:
        dict: Contains capacity info, vacancy details, and status
    """
    total_capacity_used = 0.0
    vacancy_details = []

    for vacancy in vacancies:
        vacancy_capacity = calculate_vacancy_capacity(
            vacancy['role_type'],
            vacancy.get('is_internal', False),
            vacancy.get('stage', '')
        )

        total_capacity_used += vacancy_capacity

        # Store details for display
        vacancy_details.append({
            'name': vacancy.get('vacancy_name', 'Unnamed'),
            'role_type': vacancy['role_type'].capitalize(),
            'is_internal': vacancy.get('is_internal', False),
            'stage': vacancy.get('stage', 'None'),
            'capacity_percentage': round(vacancy_capacity * 100, 2)
        })

    capacity_percentage = round(total_capacity_used * 100, 1)

    # Determine status
    if total_capacity_used > 1.0:
        status = 'overloaded'
        status_text = 'Overloaded'
    elif total_capacity_used >= 0.9:
        status = 'at-capacity'
        status_text = 'At Capacity'
    elif total_capacity_used >= 0.7:
        status = 'near-capacity'
        status_text = 'Near Capacity'
    else:
        status = 'available'
        status_text = 'Available'

    # Calculate remaining capacity
    remaining_capacity = 1.0 - total_capacity_used

    if remaining_capacity >= 0:
        additional_easy = max(0, int(remaining_capacity * 30))
        additional_medium = max(0, int(remaining_capacity * 20))
        additional_hard = max(0, int(remaining_capacity * 12))
        remaining_message = f"Can take {additional_easy} more easy OR {additional_medium} more medium OR {additional_hard} more hard vacancies"
    else:
        overload = abs(remaining_capacity)
        overload_easy = int(overload * 30)
        overload_medium = int(overload * 20)
        overload_hard = int(overload * 12)
        remaining_message = f"Overloaded by {overload_easy} easy OR {overload_medium} medium OR {overload_hard} hard vacancies"

    return {
        'capacity_used': total_capacity_used,
        'capacity_percentage': capacity_percentage,
        'status': status,
        'status_text': status_text,
        'remaining_message': remaining_message,
        'remaining_capacity': remaining_capacity,
        'vacancies': vacancy_details
    }

def calculate_recruiter_capacity(easy_vacancies, medium_vacancies, hard_vacancies):
    """
    Calculate capacity usage for a recruiter based on vacancy counts.

    Business rules:
    - Easy vacancies: Max 30 at full capacity
    - Medium vacancies: Max 20 at full capacity
    - Hard vacancies: Max 12 at full capacity

    Args:
        easy_vacancies (int): Number of easy vacancies
        medium_vacancies (int): Number of medium vacancies
        hard_vacancies (int): Number of hard vacancies

    Returns:
        dict: Contains capacity_used (0-1+), capacity_percentage, status, remaining capacities
    """
    # Calculate capacity used (can exceed 1.0 if overloaded)
    capacity_used = (easy_vacancies / 30) + (medium_vacancies / 20) + (hard_vacancies / 12)
    capacity_percentage = round(capacity_used * 100, 1)

    # Determine status based on capacity
    if capacity_used > 1.0:
        status = 'overloaded'
        status_text = 'Overloaded'
    elif capacity_used >= 0.9:
        status = 'at-capacity'
        status_text = 'At Capacity'
    elif capacity_used >= 0.7:
        status = 'near-capacity'
        status_text = 'Near Capacity'
    else:
        status = 'available'
        status_text = 'Available'

    # Calculate remaining capacity for additional work
    remaining_capacity = 1.0 - capacity_used

    if remaining_capacity >= 0:
        additional_easy = max(0, int(remaining_capacity * 30))
        additional_medium = max(0, int(remaining_capacity * 20))
        additional_hard = max(0, int(remaining_capacity * 12))
        remaining_message = f"Can take {additional_easy} more easy OR {additional_medium} more medium OR {additional_hard} more hard vacancies"
    else:
        # Overloaded - calculate how much over
        overload = abs(remaining_capacity)
        overload_easy = int(overload * 30)
        overload_medium = int(overload * 20)
        overload_hard = int(overload * 12)
        remaining_message = f"Overloaded by {overload_easy} easy OR {overload_medium} medium OR {overload_hard} hard vacancies"

    return {
        'capacity_used': capacity_used,
        'capacity_percentage': capacity_percentage,
        'status': status,
        'status_text': status_text,
        'remaining_message': remaining_message,
        'remaining_capacity': remaining_capacity
    }

def calculate_team_summary(recruiters_data):
    """
    Calculate team-wide summary statistics.

    Args:
        recruiters_data (list): List of recruiter dictionaries with capacity info

    Returns:
        dict: Team summary with counts and averages
    """
    if not recruiters_data:
        return None

    total_recruiters = len(recruiters_data)
    total_capacity = sum(r['capacity_percentage'] for r in recruiters_data)
    average_capacity = round(total_capacity / total_recruiters, 1)

    # Count by status
    status_counts = {
        'available': sum(1 for r in recruiters_data if r['status'] == 'available'),
        'near-capacity': sum(1 for r in recruiters_data if r['status'] == 'near-capacity'),
        'at-capacity': sum(1 for r in recruiters_data if r['status'] == 'at-capacity'),
        'overloaded': sum(1 for r in recruiters_data if r['status'] == 'overloaded')
    }

    # Determine overall team health
    if status_counts['overloaded'] > total_recruiters * 0.3:
        team_health = 'critical'
        team_health_text = 'Critical - Team Overloaded'
    elif status_counts['at-capacity'] + status_counts['overloaded'] > total_recruiters * 0.5:
        team_health = 'warning'
        team_health_text = 'Warning - High Utilization'
    elif average_capacity < 50:
        team_health = 'underutilized'
        team_health_text = 'Good - Capacity Available'
    else:
        team_health = 'healthy'
        team_health_text = 'Healthy - Balanced Load'

    return {
        'total_recruiters': total_recruiters,
        'average_capacity': average_capacity,
        'status_counts': status_counts,
        'team_health': team_health,
        'team_health_text': team_health_text
    }

def process_excel_upload(file):
    """
    Process uploaded Excel file and extract vacancy data.

    Expected columns:
    1. Vacancy Name
    2. Recruiter Name
    3. Role Type (Easy/Medium/Hard)
    4. Internal? (Yes/No)
    5. Stage (Sourcing/Screening/Interview/Offer/Pre-Hire Checks or blank)

    Args:
        file: FileStorage object from Flask request.files

    Returns:
        tuple: (recruiters_dict, errors_list)
               recruiters_dict: {recruiter_name: [vacancy_dicts]}
               errors_list: List of error messages
    """
    import openpyxl
    from io import BytesIO

    errors = []
    recruiters_dict = {}

    try:
        # Read Excel file
        file_bytes = file.read()
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        sheet = workbook.active

        # Check if file has data
        if sheet.max_row < 2:
            errors.append("Excel file is empty or has no data rows.")
            return recruiters_dict, errors

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        # Validate required columns (case-insensitive)
        required_columns = ['vacancy name', 'recruiter name', 'role type', 'internal?', 'stage']
        headers_lower = [str(h).lower().strip() if h else '' for h in headers]

        missing_columns = []
        for req_col in required_columns:
            if req_col not in headers_lower:
                missing_columns.append(req_col)

        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")
            return recruiters_dict, errors

        # Get column indices
        col_indices = {}
        for req_col in required_columns:
            try:
                col_indices[req_col] = headers_lower.index(req_col)
            except ValueError:
                pass

        # Process data rows
        for row_num in range(2, sheet.max_row + 1):
            row = sheet[row_num]

            # Extract values
            vacancy_name = row[col_indices['vacancy name']].value
            recruiter_name = row[col_indices['recruiter name']].value
            role_type = row[col_indices['role type']].value
            internal_str = row[col_indices['internal?']].value
            stage = row[col_indices['stage']].value

            # Validate recruiter name
            if not recruiter_name or str(recruiter_name).strip() == '':
                errors.append(f"Row {row_num}: Empty Recruiter Name")
                continue

            recruiter_name = str(recruiter_name).strip()

            # Validate vacancy name
            if not vacancy_name or str(vacancy_name).strip() == '':
                vacancy_name = f"Vacancy {row_num - 1}"

            vacancy_name = str(vacancy_name).strip()

            # Validate role type
            if not role_type or str(role_type).strip() == '':
                errors.append(f"Row {row_num}: Empty Role Type for {vacancy_name}")
                continue

            role_type = str(role_type).strip().lower()
            if role_type not in ['easy', 'medium', 'hard']:
                errors.append(f"Row {row_num}: Invalid Role Type '{role_type}' for {vacancy_name}. Must be Easy, Medium, or Hard.")
                continue

            # Validate internal
            if not internal_str or str(internal_str).strip() == '':
                internal_str = 'No'

            internal_str = str(internal_str).strip().lower()
            if internal_str not in ['yes', 'no']:
                errors.append(f"Row {row_num}: Invalid Internal value '{internal_str}' for {vacancy_name}. Must be Yes or No.")
                continue

            is_internal = (internal_str == 'yes')

            # Validate stage
            if not stage or str(stage).strip() == '':
                stage = ''
            else:
                stage = str(stage).strip().lower()
                valid_stages = ['sourcing', 'screening', 'interview', 'offer', 'pre-hire checks', '']
                if stage not in valid_stages:
                    errors.append(f"Row {row_num}: Invalid Stage '{stage}' for {vacancy_name}.")
                    continue

            # Add to recruiters dict
            if recruiter_name not in recruiters_dict:
                recruiters_dict[recruiter_name] = []

            recruiters_dict[recruiter_name].append({
                'vacancy_name': vacancy_name,
                'role_type': role_type,
                'is_internal': is_internal,
                'stage': stage
            })

    except openpyxl.utils.exceptions.InvalidFileException:
        errors.append("Invalid file format. Please upload a valid Excel file (.xlsx or .xls).")
    except Exception as e:
        errors.append(f"Error processing Excel file: {str(e)}")

    return recruiters_dict, errors

app.jinja_env.globals["format_date"] = format_date

@app.route("/")
def home():
    return render_template("index.html", latest_article=get_latest_article(), categories=BLOG_CATEGORIES, projects=PROJECTS)

@app.route("/blog")
def blog_home():
    # Pass BlogCategory objects directly - they have article_count property
    categories = list(BLOG_CATEGORIES.values())
    return render_template("blog_home.html", categories=categories)

@app.route("/blog/<category_id>")
def blog_category(category_id):
    category, articles = get_category_articles(category_id)
    if not category:
        abort(404)
    return render_template("blog_category.html", category=category, articles=articles, total_parts=len(articles))

@app.route("/blog/<category_id>/<article_id>")
def article(category_id, article_id):
    category, article_data = get_article_by_id(category_id, article_id)
    if not category or not article_data:
        abort(404)
    content = get_article_content(article_data.filename)
    if not content:
        abort(404)
    prev_article, next_article = get_prev_next_articles(category_id, article_data.part)
    return render_template("article.html", article=article_data, category=category, content_blocks=parse_content(content), reading_time=calculate_reading_time(content), prev_article=prev_article, next_article=next_article, total_parts=len(category.articles))

@app.route("/article/<article_id>")
def article_legacy(article_id):
    for cat_id, category in BLOG_CATEGORIES.items():
        for article in category.articles:
            if article.id == article_id:
                return redirect(url_for('article', category_id=cat_id, article_id=article_id))
    abort(404)

@app.route("/projects")
def projects():
    return render_template("projects.html", projects=PROJECTS)

@app.route("/projects/capacity-tracker", methods=["GET", "POST"])
def capacity_tracker():
    """
    Recruitment Capacity Tracker tool.

    GET: Display the input form
    POST: Process form data (manual or Excel upload) and display results
    """
    recruiters_data = []
    team_summary = None
    errors = []
    input_method = None

    if request.method == "POST":
        # Check if this is an Excel upload or manual input
        if 'excel_file' in request.files and request.files['excel_file'].filename:
            # Excel upload processing
            input_method = 'excel'
            file = request.files['excel_file']

            # Validate file using comprehensive validation
            is_valid, error_msg = validate_uploaded_file(file)
            if not is_valid:
                errors.append(error_msg)
            else:
                # Process Excel file
                recruiters_dict, excel_errors = process_excel_upload(file)
                errors.extend(excel_errors)

                # Calculate capacity for each recruiter
                if not errors:
                    for recruiter_name, vacancies in recruiters_dict.items():
                        try:
                            capacity_info = calculate_recruiter_capacity_from_vacancies(vacancies)
                            recruiter = {
                                'name': recruiter_name,
                                **capacity_info
                            }
                            recruiters_data.append(recruiter)
                        except Exception as e:
                            errors.append(f"Error calculating capacity for {recruiter_name}: {str(e)}")

        else:
            # Manual input processing - Enhanced with new fields
            input_method = 'manual'

            # Collect all vacancy data from form
            # Form fields are named: recruiter_0, vacancy_name_0, role_type_0, internal_0, stage_0, etc.
            index = 0
            vacancies_by_recruiter = {}

            while f'recruiter_{index}' in request.form:
                recruiter_name = request.form.get(f'recruiter_{index}', '').strip()
                vacancy_name = request.form.get(f'vacancy_name_{index}', '').strip()
                role_type = request.form.get(f'role_type_{index}', '').strip().lower()
                internal_str = request.form.get(f'internal_{index}', 'no').strip().lower()
                stage = request.form.get(f'stage_{index}', '').strip().lower()

                # Skip if recruiter name is empty
                if not recruiter_name:
                    index += 1
                    continue

                # Validate role type
                if role_type not in ['easy', 'medium', 'hard']:
                    errors.append(f"Invalid role type for vacancy {index + 1}")
                    index += 1
                    continue

                # Parse internal
                is_internal = (internal_str == 'yes')

                # Default vacancy name if not provided
                if not vacancy_name:
                    vacancy_name = f"Vacancy {index + 1}"

                # Add to recruiter's vacancy list
                if recruiter_name not in vacancies_by_recruiter:
                    vacancies_by_recruiter[recruiter_name] = []

                vacancies_by_recruiter[recruiter_name].append({
                    'vacancy_name': vacancy_name,
                    'role_type': role_type,
                    'is_internal': is_internal,
                    'stage': stage
                })

                index += 1

            # Calculate capacity for each recruiter
            if not errors and vacancies_by_recruiter:
                for recruiter_name, vacancies in vacancies_by_recruiter.items():
                    try:
                        capacity_info = calculate_recruiter_capacity_from_vacancies(vacancies)
                        recruiter = {
                            'name': recruiter_name,
                            **capacity_info
                        }
                        recruiters_data.append(recruiter)
                    except Exception as e:
                        errors.append(f"Error calculating capacity for {recruiter_name}: {str(e)}")

        # Calculate team summary if we have data
        if recruiters_data:
            team_summary = calculate_team_summary(recruiters_data)

    return render_template(
        "projects/capacity_tracker.html",
        recruiters=recruiters_data,
        team_summary=team_summary,
        errors=errors,
        input_method=input_method
    )

@app.route("/projects/capacity-tracker/download-template")
def download_capacity_template():
    """
    Generate and download a sample Excel template for capacity tracker.
    """
    import openpyxl
    from io import BytesIO
    from flask import send_file

    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Capacity Data"

    # Add headers
    headers = ['Vacancy Name', 'Recruiter Name', 'Role Type', 'Internal?', 'Stage']
    sheet.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add sample data
    sample_data = [
        ['Senior Developer', 'John Smith', 'Hard', 'No', 'Screening'],
        ['Marketing Assistant', 'Jane Doe', 'Easy', 'Yes', 'Sourcing'],
        ['Finance Manager', 'John Smith', 'Hard', 'No', ''],
        ['HR Coordinator', 'Jane Doe', 'Medium', 'No', 'Interview'],
        ['IT Support', 'Bob Wilson', 'Easy', 'Yes', 'Offer'],
    ]

    for row in sample_data:
        sheet.append(row)

    # Adjust column widths
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 18

    # Add instructions sheet
    instructions_sheet = workbook.create_sheet("Instructions")
    instructions = [
        ['Recruitment Capacity Tracker - Template Instructions'],
        [''],
        ['Column Requirements:'],
        ['1. Vacancy Name: Name of the vacancy (optional, will auto-generate if blank)'],
        ['2. Recruiter Name: Name of the recruiter (required)'],
        ['3. Role Type: Easy, Medium, or Hard (required, case-insensitive)'],
        ['4. Internal?: Yes or No (required, case-insensitive)'],
        ['5. Stage: Sourcing, Screening, Interview, Offer, Pre-Hire Checks, or blank (optional)'],
        [''],
        ['Business Rules:'],
        ['- Easy roles: Max 30 at full capacity (3.33% each)'],
        ['- Medium roles: Max 20 at full capacity (5% each)'],
        ['- Hard roles: Max 12 at full capacity (8.33% each)'],
        ['- Internal roles take 75% less time (0.25 multiplier)'],
        ['- Stage weights: Sourcing (20%), Screening (40%), Interview (20%), Offer (10%), Pre-Hire (10%), None (100%)'],
        [''],
        ['Example Calculations:'],
        ['- External, Easy, No Stage: 1/30 = 3.33%'],
        ['- Internal, Hard, Screening: (1/12) × 0.25 × 0.4 = 0.83%'],
        ['- External, Medium, Interview: (1/20) × 0.2 = 1%'],
    ]

    for row in instructions:
        instructions_sheet.append(row)

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='capacity_tracker_template.xlsx'
    )

@app.route("/about")
def about():
    return render_template("about.html")

@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404

if __name__ == "__main__":
    # Get configuration from app config (already loaded)
    debug_mode = app.config.get('DEBUG', False)
    env_name = os.environ.get('FLASK_ENV', 'development')

    # Display startup information
    print("=" * 60)
    print(f"Flask Application Starting")
    print(f"Environment: {env_name}")
    print(f"Debug Mode: {debug_mode}")
    print(f"Config: {app.config.__class__.__name__}")
    print("=" * 60)

    if debug_mode and env_name == 'production':
        print("\n⚠️  WARNING: Debug mode enabled in production!")
        print("This is a security risk. Set FLASK_DEBUG=false\n")

    # Get host and port from environment or use defaults
    host = os.environ.get('FLASK_HOST', '127.0.0.1')
    port = int(os.environ.get('FLASK_PORT', 5000))

    app.run(host=host, port=port, debug=debug_mode)
